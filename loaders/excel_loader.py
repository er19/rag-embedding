from openpyxl import load_workbook

from .base import BaseLoader, Document


class ExcelLoader(BaseLoader):
    def load(self, file_path: str) -> list[Document]:
        documents = []
        wb = load_workbook(file_path, read_only=True)
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            rows = []
            for row in ws.iter_rows(values_only=True):
                cells = [str(c) if c is not None else "" for c in row]
                rows.append(" | ".join(cells))
            text = "\n".join(rows)
            if text.strip():
                documents.append(
                    Document(
                        content=text,
                        metadata={"source": str(wb.path), "sheet": sheet_name},
                    )
                )
        wb.close()
        return documents
